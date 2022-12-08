import time
from selenium_pages import UseSelenium
from selenium_product import UseSelenium as UseSelenium_2
from bs4 import BeautifulSoup
import os
import json
import re
import pandas as pd
from lib.csv_handler import CsvHandler
import requests
import openpyxl.drawing.image
from PIL import Image as pilIm
from openpyxl import load_workbook
import shutil
import glob

from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )


global num_tovarov,class_pars,filename_final,image_path
start_time = time.time()

url = 'https://www.ozon.ru/category/zerkala-15051/'
class_in_page = 'widget-search-result-container ku5'
filename_final = r'C:\PyProjects\poject_1\tural\ozon\ozon_дом_зеркала.xlsx'
image_path = r'C:\PyProjects\poject_1\tural\ozon\images_2\phone'
class_pars = ['k5u','k8t']
num_tovarov = 10

def download_category_html(url):

    # Ограничим парсинг первыми 10 страницами
    MAX_PAGE = 1
    i = 1
    while i <= MAX_PAGE:
        filename = f'page_' + str(i) + '.html'
        if i == 1:
            UseSelenium(url, filename).save_page()
        else:
            url_param = url + '?page=' + str(i)
            UseSelenium(url_param, filename).save_page()

        i += 1

def get_pages() -> list:
    return glob.glob('C:/PyProjects/poject_1/tural/ozon/pages/*.html')


def get_html(page: str):
    with open(page, 'r', encoding='utf-8') as f:
        return f.read()


def parse_data(html: str) -> str:
    soup = BeautifulSoup(html, 'html.parser')

    try:
        links = []
        products = soup.find('div', attrs={'class', 'k5u'}).find_all('a')  # k7r
        for product in products:
            links.append(product.get('href').split('?')[0])
    except:
        print('НЕ t6k')
        links = []
        products = soup.find('div', attrs={'class', 'k8t'}).find_all('a')  # k7r
        for product in products:
            links.append(product.get('href').split('?')[0])

        # print('НЕ t6k, k8t')

    #return set(links)
    final_links = []
    min_len = len(links[0])/2
    for link in links:
        if (link not in final_links) and (len(link) > min_len):
            final_links.append(link)

    return final_links[:10]


def get_product_links() -> list:
    with open(r'C:\PyProjects\poject_1\tural\ozon\product_links.txt', 'r', encoding='utf-8') as f:
        return f.readlines()

def data_parsing(product: str, i: int, filename: str) -> None:
    url = 'https://www.ozon.ru/api/composer-api.bx/page/json/v2' \
          f'?url={product}'

    filename = filename + str(i) + '.html'
    UseSelenium_2(url, filename).save_page()

def save_lincs_to_txt(all_links):
    with open(os.getcwd()+'\product_links.txt', 'w', encoding='utf-8') as f:
        for link in all_links:
            f.write(link + '\n')


def get_products() -> list:
    return glob.glob('products/*.html')

def get_json(filename_final: str) -> dict:
    with open(filename_final, 'r', encoding='utf-8') as f:
        data = f.read()
        return json.loads(data)

def parse_data_product(data: dict) -> dict:
    widgets = data.get('widgetStates')
    for key, value in widgets.items():
        if 'webProductHeading' in key:
            title = json.loads(value).get('title')
        if 'webSale' in key:
            prices = json.loads(value).get('offers')[0]
            if prices.get('price'):
                price = re.search(r'[0-9]+', prices.get('price').replace(u'\u2009', ''))[0]
            else:
                price = 0
            if prices.get('originalPrice'):
                discount_price = re.search(r'[0-9]+', prices.get('originalPrice').replace(u'\u2009', ''))[0]
            else:
                discount_price = 0
        if 'webGallery' in key:
            url_pic = json.loads(value).get('coverImage')

    layout = json.loads(data.get('layoutTrackingInfo'))
    brand = layout.get('brandName')
    category = layout.get('categoryName')
    sku = layout.get('sku')
    url = layout.get('currentPageUrl')


    product = {
        'img': '',
        'title': title,
        'price': price,
        'discount_price': discount_price,
        'brand': brand,
        'category': category,
        'sku': sku,
        'url': url[:-1],
        'url_pic': url_pic
    }
    return product


def download_img():
    df_img = pd.read_excel(filename_final)
    img_list = df_img['url_pic'].to_list()
    j = 0
    for i in img_list:
        j += 1
        img_data = requests.get(i).content
        with open(image_path+'\image_name_'+str(j)+'.jpg', 'wb') as handler:
            handler.write(img_data)


def change_size(img, shirina):
    k = img.width/shirina
    img.width = shirina
    img.height = img.height/k
    return img

def img_to_excel():
    path = os.getcwd()
    shirina = 150
    kachestvo = 45

    os.mkdir("bad_images")
    spis = os.listdir(path+'\\'+'images_2')

    for papka in spis:
        os.mkdir('bad_images\\'+papka)
        files = os.listdir(path+'\\'+'images_2\\'+papka)

    for i in files:
        try:
            image = pilIm.open(path+'\\'+'images_2\\'+papka+'\\'+i)
            image.save(path+'\\'+'bad_images\\'+papka+'\\'+i, quality = kachestvo)
        except:
            print('Ошибка с картинкой')
            return 1

    spis = os.listdir(path+'\\'+'bad_images')
    for papka in spis:
        files = os.listdir('bad_images\\'+papka)

        #сортировка названий изображений
        def numerical_sort(value: str) -> list:
            parts = re.compile(r'(\d+)').split(value)
            parts[1::2] = map(int, parts[1::2])
            return parts

        files = sorted(files, key=numerical_sort)

        wb = load_workbook(filename_final)
        sheet = wb.active

        for i in range(len(files)):
            img = openpyxl.drawing.image.Image('bad_images\\'+papka+'\\'+files[i])
            img = change_size(img, shirina)
            sheet.add_image(img, 'A%d'%(i+2))

            sheet.row_dimensions[i+2].height = img.height*0.75
        sheet.column_dimensions['A'].width = img.width*0.143
        print(files)
        wb.save(filename_final)

    # for papka in spis:
    #     files = os.listdir('bad_images\\' + papka)
    #     for i in range(len(files)):
    #         os.remove('bad_images\\'+papka+'\\'+files[i])
    #     os.rmdir('bad_images\\'+papka)
    # os.rmdir("bad_images")
    return 0

def delete_pictures():
    files = glob.glob(image_path+'/*')
    for f in files:
        os.remove(f)

def format_excel():
    wb = load_workbook(filename_final)
    sheet = wb.active
    sheet.row_dimensions[1].height = 20
    # ffcc00 - оранжевый
    sheet['A1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['B1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['C1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['D1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['E1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['F1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['G1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['H1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['I1'].fill = PatternFill('solid', fgColor='ffcc00')
    #sheet['J1'].fill = PatternFill('solid', fgColor='ffcc00')
    #sheet['K1'].fill = PatternFill('solid', fgColor='ffcc00')

    for row in sheet[2:sheet.max_row]:  # skip the header

        row[1].alignment = Alignment(horizontal="left", vertical="center")
        row[2].alignment = Alignment(horizontal="center", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="left", vertical="center")
        row[8].alignment = Alignment(horizontal="left", vertical="center")
        #row[9].alignment = Alignment(horizontal="left", vertical="center")
        #row[10].alignment = Alignment(horizontal="center", vertical="center")

    for i in range(sheet.max_row):
        sheet.cell(row = i+1, column=10).value = ' '

    a = 20
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = a
    sheet.column_dimensions['D'].width = a
    sheet.column_dimensions['E'].width = a
    sheet.column_dimensions['F'].width = a
    sheet.column_dimensions['G'].width = a
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 10

    wb.save(filename_final)

def del_file_if_exsict():
    if os.path.isfile(filename_final):
        os.remove(filename_final)
    else:
        pass

def del_bad_images():
    try:
        shutil.rmtree(os.getcwd()+'/bad_images')
    except:
        pass

if __name__ == '__main__':
    download_category_html(url)
    pages = get_pages()

    all_links = []
    for page in pages:
        html = get_html(page)
        links = parse_data(html)
        links = links[:num_tovarov]
        all_links = all_links + list(links)

    save_lincs_to_txt(all_links)

    products = get_product_links()

    i = 1
    for product in products:
        data_parsing(product, i, filename='product_')
        i += 1

    CsvHandler(filename_final).create_headers_csv_semicolon(
        ['img','title', 'price', 'discount_price', 'brand', 'category', 'sku', 'url', 'url_pic'])
    products = get_products()

    df_all = pd.DataFrame(columns=['img','title', 'price', 'discount_price', 'brand', 'category', 'sku', 'url', 'url_pic'])

    i = 0
    for product in products:
        try:
            product_json = get_json(product)
            result = parse_data_product(product_json)
            df_all.loc[i] = list(result.values())
        except Exception as e:
            print(e)
        i += 1


    print(df_all)
    del_file_if_exsict()
    df_all.to_excel(filename_final, index=False)

    del_bad_images()
    download_img()
    img_to_excel()
    delete_pictures()
    del_bad_images()
    #delete_pictures()


    format_excel()


    print("--- %s seconds ---" % (time.time() - start_time))