from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )

#filename = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_data_category.xlsx'
filename = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_data_mass.xlsx'

def format_excel(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.row_dimensions[1].height = 20
    # ffcc00
    sheet['A1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['B1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['C1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['D1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['E1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['F1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['G1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['H1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['I1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['J1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['K1'].fill = PatternFill('solid', fgColor='ffcc00')


    for row in sheet[2:sheet.max_row]:  # skip the header

        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
        row[8].alignment = Alignment(horizontal="left", vertical="center")
        row[9].alignment = Alignment(horizontal="left", vertical="center")
        row[10].alignment = Alignment(horizontal="center", vertical="center")

    a = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = a
    sheet.column_dimensions['E'].width = a
    sheet.column_dimensions['F'].width = a
    sheet.column_dimensions['G'].width = a
    sheet.column_dimensions['H'].width = a

    wb.save(filename)

format_excel(filename)