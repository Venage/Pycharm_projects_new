import requests
import json
import re
import os.path
import random
import pandas as pd
from bs4 import BeautifulSoup
from seleniumwire import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl.drawing.image
from PIL import Image as pilIm
from openpyxl import load_workbook
import glob
import lib.headers
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
from lib.csv_handler import CsvHandler




global url, filename, max_tovarov, image_path
url = 'https://www.wildberries.ru/catalog/dom/zerkala'
filename = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_дом_зеркала.xlsx'
image_path = r'C:\PyProjects\poject_1\tural\new_wb_pars\images_2\phone'

max_tovarov = 10
low_price = 0
top_price = 100000



"""
Парсер wildberries по ссылке на каталог (указывать без фильтров)
Парсер не идеален, есть множество вариантов реализации, со своими идеями 
и предложениями обязательно пишите мне, либо в группу, ссылка ниже.
Подробное описание парсера Вайлдберриз можно почитать на сайте:
https://happypython.ru/2022/07/21/парсер-wildberries/
Ссылка на статью ВКонтакте: https://vk.com/@happython-parser-wildberries
По всем возникшим вопросам, можете писать в группу https://vk.com/happython
парсер wildberries по каталогам 2022, обновлен 19.10.2022 - на данное число работает исправно
"""


def get_catalogs_wb():
    """получение каталога вб"""
    url = 'https://www.wildberries.ru/webapi/menu/main-menu-ru-ru.json'
    headers = {'Accept': "*/*", 'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    response = requests.get(url, headers=headers)
    data = response.json()
    with open('pages/wb_catalogs_data.json', 'w', encoding='UTF-8') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
        print(f'Данные сохранены в wb_catalogs_data_sample.json')
    data_list = []
    for d in data:
        try:
            for child in d['childs']:
                try:
                    category_name = child['name']
                    category_url = child['url']
                    shard = child['shard']
                    query = child['query']
                    data_list.append({
                        'category_name': category_name,
                        'category_url': category_url,
                        'shard': shard,
                        'query': query})
                except:
                    continue
                try:
                    for sub_child in child['childs']:
                        category_name = sub_child['name']
                        category_url = sub_child['url']
                        shard = sub_child['shard']
                        query = sub_child['query']
                        data_list.append({
                            'category_name': category_name,
                            'category_url': category_url,
                            'shard': shard,
                            'query': query})
                except:
                    # print(f'не имеет дочерних каталогов *{i["name"]}*')
                    continue
        except:
            # print(f'не имеет дочерних каталогов *{d["name"]}*')
            continue
    return data_list


def search_category_in_catalog(url, catalog_list):
    """пишем проверку пользовательской ссылки на наличии в каталоге"""
    try:
        for catalog in catalog_list:
            if catalog['category_url'] == url.split('https://www.wildberries.ru')[-1]:
                print(f'найдено совпадение: {catalog["category_name"]}')
                name_category = catalog['category_name']
                shard = catalog['shard']
                query = catalog['query']
                return name_category, shard, query
            else:
                # print('нет совпадения')
                pass
    except:
        print('Данный раздел не найден!')


def get_data_from_json(json_file):
    """извлекаем из json интересующие нас данные"""
    data_list = []
    for data in json_file['data']['products']:
        data_list.append({
            'url_tovara': f'https://www.wildberries.ru/catalog/{data["id"]}/detail.aspx?targetUrl=BP'
        })
    return data_list


def get_content(shard, query, low_price=None, top_price=None):
    # вставляем ценовые рамки для уменьшения выдачи, вилбериес отдает только 100 страниц
    headers = {'Accept': "*/*", 'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    data_list = []
    for page in range(1, 2):
        print(f'Сбор позиций со страницы {page} из 100')
        # url = f'https://wbxcatalog-ru.wildberries.ru/{shard}' \
        #       f'/catalog?appType=1&curr=rub&dest=-1029256,-102269,-1278703,-1255563' \
        #       f'&{query}&lang=ru&locale=ru&sort=sale&page={page}' \
        #       f'&priceU={low_price * 100};{top_price * 100}'
        url = f'https://catalog.wb.ru/catalog/{shard}/catalog?appType=1&curr=rub&dest=-1075831,-77677,-398551,12358499' \
              f'&locale=ru&page={page}&priceU={low_price * 100};{top_price * 100}' \
              f'&reg=0&regions=64,83,4,38,80,33,70,82,86,30,69,1,48,22,66,31,40&sort=popular&spp=0&{query}'
        r = requests.get(url, headers=headers)
        data = r.json()
        #print(f'Добавлено позиций: {len(get_data_from_json(data))}')
        if len(get_data_from_json(data)) > 0:
            data_list.extend(get_data_from_json(data))
        else:
            print(f'Сбор данных завершен.')
            break
    return data_list


def save_excel(data, filename):
    """сохранение результата в excel файл"""
    df = pd.DataFrame(data)
    writer = pd.ExcelWriter(f'{filename}.xlsx')
    df.to_excel(writer, 'data')
    writer.save()
    print(f'Все сохранено в {filename}.xlsx')


def parser(url, low_price, top_price):
    # получаем список каталогов
    catalog_list = get_catalogs_wb()
    try:
        # поиск введенной категории в общем каталоге
        name_category, shard, query = search_category_in_catalog(url=url, catalog_list=catalog_list)
        # сбор данных в найденном каталоге
        data_list = get_content(shard=shard, query=query, low_price=low_price, top_price=top_price)
        # сохранение найденных данных
        data_list = data_list[:max_tovarov]
        df = pd.DataFrame(data_list)
        url_list = df['url_tovara'].to_list()
        return (url_list, name_category)
        # save_excel(data_list, f'{name_category}_from_{low_price}_to_{top_price}')
    except TypeError:
        print('Ошибка! Возможно не верно указан раздел. Удалите все доп фильтры с ссылки')
    except PermissionError:
        print('Ошибка! Вы забыли закрыть созданный ранее excel файл. Закройте и повторите попытку')



## Часть 2





def get_headers_proxy() -> dict:
    '''
    The config file must have dict:
        {
            'http_proxy':'http://user:password@ip:port',
            'user-agent': 'user_agent name'
        }
    '''

    try:
        """
        users = lib.config.USER_AGENTS_PROXY_LIST
        persona = random.choice(users)
        """

        users = [{'http_proxy': 'http://YzY8Vu:UQpB3Z@217.29.62.212:11045',
                  'user-agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:94.0) Gecko/20100101 Firefox/94.0'}]
        persona = random.choice(users)

    except ImportError:
        persona = None
    return persona

def get_html(persona, url):
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-agent={persona['user-agent']}")
    options.add_argument("--disable-blink-features=AutomationControlled")

    options_proxy = {
        'proxy': {
            'http': persona['http_proxy']
            # 'no_proxy': 'localhost,127.0.0.1:8080'
        }
    }

    s = Service(executable_path="C:/Users/venag/Downloads/chromedriver_win32/chromedriver.exe")
    # driver = webdriver.Chrome(options=options, service=s, seleniumwire_options=options_proxy)

    proxies = {
        # 'https': 'http://128.69.161.246'
        'http': 'http://YzY8Vu:UQpB3Z@217.29.62.212:11045'
    }
    driver = webdriver.Chrome(options=options, service=s)

    try:
        driver.request_interceptor = lib.headers.interceptor
        driver.get(url)
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "collapsible__toggle-wrap")))

        driver.find_element(By.XPATH, "//*[contains(text(), 'Развернуть характеристики')]").click()
        wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Свернуть характеристики')]")))
        html = driver.page_source
    except Exception as ex:
        print(ex)
        html = None
    finally:
        driver.close()
        driver.quit()
    return html

def parse_data(html,url,name_category):
    soup = BeautifulSoup(html, 'html.parser')
    h1 = get_h1(soup)
    sku = get_sku(soup)
    stars = get_stars(soup)
    price = get_price(soup)
    old_price = get_old_price(soup)
    description = get_description(soup)
    brand = get_brand(soup)

    url_img = get_img_url(soup)
    url_img = 'https:'+url_img

    tables = get_tables_specifications(soup)
    specifications = prepare_specifications(tables) if tables else None

    price = re.findall(r'[0-9]+', re.sub(r'\xa0', '', price))[0] if price else None
    old_price = re.findall(r'[0-9]+', re.sub(u'\xa0', '', old_price))[0] if old_price else None
    payload = {
        'title': h1,
        'sku': sku,
        'price': price,
        'old_price': old_price,
        'description': description,
        'stars': stars,
        'brand': brand,
        'specifications': specifications,
    }
    sale_percent =  (int(old_price) - int(price))/int(old_price)*100
    sale_percent = round(sale_percent,2)


    #print(h1, sku, price, old_price, description, stars, brand, specifications)
    #print(h1, sku, price, old_price, stars, brand)

    df = pd.DataFrame(columns=['Изображение','Артикул', 'Название', 'Категория', 'Бренд', 'Цена до скидки', 'Цена со скидкой', 'Размер скидки %', 'url', 'url_img', 'stars'])
    row = ('',sku,h1, name_category, brand, old_price, price, sale_percent, url, url_img, stars)
    df.loc[0] = list(row)
    print(df)
    print(row)
    write_to_csv(row,df)



def write_to_csv(row,df):
    #filename = 'wildberries_data.csv'
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        sheet = wb.active
        sheet.append(row)
        wb.save(filename)
    else:
        df.to_excel(filename, index = False)


# Декоратор для перехвата ошибок
def try_except(func):
    def wrapper(*args, **kwargs):
        try:
            data = func(*args, **kwargs)
        except:
            data = None
        return data
    return wrapper

@try_except
def get_h1(soup):
    return soup.h1.string

@try_except
def get_sku(soup):
    return soup.find('span', attrs={'id': 'productNmId'}).get_text()

@try_except
def get_stars(soup):
    return soup.find('div', attrs={'class': 'product-page__common-info'}).find('span', attrs={'data-link': 'text{: product^star}'}).get_text()

@try_except
def get_price(soup):
    return soup.find('ins', attrs={'class': 'price-block__final-price'}).get_text()

@try_except
def get_old_price(soup):
    return soup.find('del', attrs={'class': 'price-block__old-price j-wba-card-item-show'}).get_text()

@try_except
def get_description(soup):
    return soup.find('p', attrs={'class': 'collapsable__text'}).get_text()

@try_except
def get_brand(soup):
    return soup.find('div', attrs={'class': 'product-page__brand-logo hide-mobile'}).find('a').get('title')

@try_except
def get_tables_specifications(soup):
    return soup.find('div', attrs={'class': 'collapsable__content j-add-info-section'}).find_all('table')

@try_except
def get_img_url(soup):
    return soup.find('div', attrs={'class': 'zoom-image-container'}).find('img')['src']

def prepare_specifications(tables):
    data_spec_all = {}
    data_spec = []
    for table in tables:
        caption = table.find('caption').get_text()
        for tr in table.find_all('tr'):
            char = tr.find('th').get_text()
            value = tr.find('td').get_text()
            data_spec.append([char.strip(), value.strip()])
        data_spec_all[caption] = data_spec

    return data_spec_all




def download_img():
    df_img = pd.read_excel(filename)
    img_list = df_img['url_img'].to_list()
    j = 0
    for i in img_list:
        j += 1
        img_data = requests.get(i).content
        with open(image_path+'\image_name_'+str(j)+'.jpg', 'wb') as handler:
            handler.write(img_data)


def change_size(img, shirina):
    k = img.width/shirina
    img.width = shirina
    img.height = img.height/k
    return img

def img_to_excel():
    path = os.getcwd()
    shirina = 150
    kachestvo = 45

    os.mkdir("bad_images")
    spis = os.listdir(path + '\\' + 'images_2')

    for papka in spis:
        os.mkdir('bad_images\\' + papka)
        files = os.listdir(path + '\\' + 'images_2\\' + papka)

    for i in files:
        try:
            image = pilIm.open(path + '\\' + 'images_2\\' + papka + '\\' + i)
            image.save(path + '\\' + 'bad_images\\' + papka + '\\' + i, quality=kachestvo)
        except:
            print('Ошибка с картинкой')
            return 1
    spis = os.listdir(path + '\\' + 'bad_images')
    for papka in spis:
        files = os.listdir('bad_images\\' + papka)
    # сортировка названий изображений
        def numerical_sort(value: str) -> list:
            parts = re.compile(r'(\d+)').split(value)
            parts[1::2] = map(int, parts[1::2])
            return parts

        files = sorted(files, key=numerical_sort)

        wb = load_workbook(filename)
        sheet = wb.active

        for i in range(len(files)):
            img = openpyxl.drawing.image.Image('bad_images\\' + papka + '\\' + files[i])
            img = change_size(img, shirina)
            sheet.add_image(img, 'A%d' % (i + 2))

            sheet.row_dimensions[i + 2].height = img.height * 0.75
        sheet.column_dimensions['A'].width = img.width * 0.143
        print(files)
        wb.save(filename)

def delete_pictures():
    files = glob.glob(image_path+'/*')
    for f in files:
        os.remove(f)

def delete_bad_pictures():
    path = os.getcwd()
    # os.mkdir("bad_images")
    spis = os.listdir(path + '\\' + 'bad_images')
    for papka in spis:
        files = os.listdir('bad_images\\' + papka)
        for i in range(len(files)):
            os.remove('bad_images\\'+papka+'\\'+files[i])
        os.rmdir('bad_images\\'+papka)
    os.rmdir("bad_images")


def main(url,name_category):
    persona = get_headers_proxy()
    html = get_html(persona, url)
    if html: parse_data(html,url,name_category)


def format_excel():
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.row_dimensions[1].height = 20
    # ffcc00 - оранжевый
    sheet['A1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['B1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['C1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['D1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['E1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['F1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['G1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['H1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['I1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['J1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['K1'].fill = PatternFill('solid', fgColor='ffcc00')

    for row in sheet[2:sheet.max_row]:  # skip the header

        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
        row[8].alignment = Alignment(horizontal="left", vertical="center")
        row[9].alignment = Alignment(horizontal="left", vertical="center")
        row[10].alignment = Alignment(horizontal="center", vertical="center")

    a = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = a
    sheet.column_dimensions['E'].width = a
    sheet.column_dimensions['F'].width = a
    sheet.column_dimensions['G'].width = a
    sheet.column_dimensions['H'].width = a

    wb.save(filename)




if __name__ == '__main__':
    all_url,name_category = parser(url, low_price, top_price)
    print(all_url)
    for url in all_url:
        main(url,name_category)
    download_img()
    img_to_excel()
    delete_pictures()
    delete_bad_pictures()
    format_excel()