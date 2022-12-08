import requests
import time
import json
import re
import os.path
import random
import pandas as pd
from bs4 import BeautifulSoup
from seleniumwire import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl.drawing.image
from PIL import Image as pilIm
from openpyxl import load_workbook
import glob
import lib.headers
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
from lib.csv_handler import CsvHandler
start_time = time.time()

all_url = ['https://www.wildberries.ru/catalog/102240137/detail.aspx?targetUrl=GP']

global filename, image_path
filename = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_data_mass.xlsx'
image_path = r'C:\PyProjects\poject_1\tural\new_wb_pars\images\phone'


def get_headers_proxy() -> dict:
    '''
    The config file must have dict:
        {
            'http_proxy':'http://user:password@ip:port',
            'user-agent': 'user_agent name'
        }
    '''

    try:
        """
        users = lib.config.USER_AGENTS_PROXY_LIST
        persona = random.choice(users)
        """

        users = [{'http_proxy': 'http://YzY8Vu:UQpB3Z@217.29.62.212:11045',
                  'user-agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:94.0) Gecko/20100101 Firefox/94.0'}]
        persona = random.choice(users)

    except ImportError:
        persona = None
    return persona

def get_html(persona, url):
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-agent={persona['user-agent']}")
    options.add_argument("--disable-blink-features=AutomationControlled")

    options_proxy = {
        'proxy': {
            'http': persona['http_proxy']
            # 'no_proxy': 'localhost,127.0.0.1:8080'
        }
    }

    s = Service(executable_path="C:/Users/venag/Downloads/chromedriver_win32/chromedriver.exe")
    # driver = webdriver.Chrome(options=options, service=s, seleniumwire_options=options_proxy)

    proxies = {
        # 'https': 'http://128.69.161.246'
        'http': 'http://LU5zmm:odarM8@217.29.62.211:10481'
    }
    driver = webdriver.Chrome(options=options, service=s)

    try:
        driver.request_interceptor = lib.headers.interceptor
        driver.get(url)
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "collapsible__toggle-wrap")))

        #driver.find_element(By.XPATH, "//*[contains(text(), 'Развернуть характеристики')]").click()
        #wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Свернуть характеристики')]")))

        html = driver.page_source
    except Exception as ex:
        print(ex)
        html = None
    finally:
        driver.close()
        driver.quit()
    return html

def parse_data(html,url):
    soup = BeautifulSoup(html, 'html.parser')
    h1 = get_h1(soup)
    sku = get_sku(soup)
    stars = get_stars(soup)
    price = get_price(soup)
    old_price = get_old_price(soup)
    description = get_description(soup)
    brand = get_brand(soup)

    # -3 бедет более широкая категория
    name_category_all = get_name_category(soup)
    try:
        name_category = name_category_all[-2].get_text()
    except:
        name_category = ''

    url_img = get_img_url(soup)
    url_img = 'https:'+url_img

    tables = get_tables_specifications(soup)
    specifications = prepare_specifications(tables) if tables else None

    price = re.findall(r'[0-9]+', re.sub(r'\xa0', '', price))[0] if price else None
    old_price = re.findall(r'[0-9]+', re.sub(u'\xa0', '', old_price))[0] if old_price else None

    payload = {
        'title': h1,
        'sku': sku,
        'price': price,
        'old_price': old_price,
        'description': description,
        'stars': stars,
        'brand': brand,
        'specifications': specifications,
        'name_category' : name_category
    }
    #print(h1, sku, price, old_price, description, stars, brand, specifications)
    #print(h1, sku, price, old_price, stars, brand)
    try:
        sale_percent = (int(old_price) - int(price)) / int(old_price) * 100
        sale_percent = round(sale_percent, 2)
    except:
        sale_percent = 0

    df = pd.DataFrame(
        columns=['Изображение', 'Артикул', 'Название', 'Категория', 'Бренд', 'Цена до скидки', 'Цена со скидкой',
                 'Размер скидки %', 'url', 'url_img', 'stars'])
    row = ('', sku, h1, name_category, brand, old_price, price, sale_percent, url, url_img, stars)
    df.loc[0] = list(row)
    print(df)
    print(row)
    write_to_csv(row,df)



def write_to_csv(row,df):
    #filename = 'wildberries_data.csv'
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        sheet = wb.active
        sheet.append(row)
        wb.save(filename)
    else:
        df.to_excel(filename, index = False)


# Декоратор для перехвата ошибок
def try_except(func):
    def wrapper(*args, **kwargs):
        try:
            data = func(*args, **kwargs)
        except:
            data = None
        return data
    return wrapper

@try_except
def get_h1(soup):
    return soup.h1.string

@try_except
def get_sku(soup):
    return soup.find('span', attrs={'id': 'productNmId'}).get_text()

@try_except
def get_stars(soup):
    return soup.find('div', attrs={'class': 'product-page__common-info'}).find('span', attrs={'data-link': 'text{: product^star}'}).get_text()

@try_except
def get_price(soup):
    return soup.find('ins', attrs={'class': 'price-block__final-price'}).get_text()

@try_except
def get_old_price(soup):
    return soup.find('del', attrs={'class': 'price-block__old-price j-wba-card-item-show'}).get_text()

@try_except
def get_description(soup):
    return soup.find('p', attrs={'class': 'collapsable__text'}).get_text()

@try_except
def get_brand(soup):
    return soup.find('div', attrs={'class': 'product-page__brand-logo hide-mobile'}).find('a').get('title')

@try_except
def get_tables_specifications(soup):
    return soup.find('div', attrs={'class': 'collapsable__content j-add-info-section'}).find_all('table')

@try_except
def get_img_url(soup):
    return soup.find('div', attrs={'class': 'zoom-image-container'}).find('img')['src']

@try_except
def get_name_category(soup):
    return soup.find_all('span', attrs={'itemprop': 'name'})
    #return soup.find('div', attrs={'class': 'breadcrumbs__container'})#.find('span', attrs={'itemprop': 'name'}).get_text()

def prepare_specifications(tables):
    data_spec_all = {}
    data_spec = []
    for table in tables:
        caption = table.find('caption').get_text()
        for tr in table.find_all('tr'):
            char = tr.find('th').get_text()
            value = tr.find('td').get_text()
            data_spec.append([char.strip(), value.strip()])
        data_spec_all[caption] = data_spec

    return data_spec_all




def download_img():
    df_img = pd.read_excel(filename)
    img_list = df_img['url_img'].to_list()
    j = 0
    for i in img_list:
        j += 1
        img_data = requests.get(i).content
        with open(image_path+'\image_name_'+str(j)+'.jpg', 'wb') as handler:
            handler.write(img_data)


def change_size(img, shirina):
    k = img.width/shirina
    img.width = shirina
    img.height = img.height/k
    return img

def img_to_excel():
    path = os.getcwd()
    shirina = 150
    kachestvo = 45

    os.mkdir("bad_images")
    spis = os.listdir(path+'\\'+'images')

    for papka in spis:
        os.mkdir('bad_images\\'+papka)
        files = os.listdir(path+'\\'+'images\\'+papka)

    for i in files:
        try:
            image = pilIm.open(path+'\\'+'images\\'+papka+'\\'+i)
            image.save(path+'\\'+'bad_images\\'+papka+'\\'+i, quality = kachestvo)
        except:
            print('Ошибка с картинкой')
            return 1

    spis = os.listdir(path+'\\'+'bad_images')
    for papka in spis:
        files = os.listdir('bad_images\\'+papka)

        wb = load_workbook(filename)
        sheet = wb.active

        for i in range(len(files)):
            img = openpyxl.drawing.image.Image('bad_images\\'+papka+'\\'+files[i])
            img = change_size(img, shirina)
            sheet.add_image(img, 'A%d'%(i+2))

            sheet.row_dimensions[i+2].height = img.height*0.75
        sheet.column_dimensions['A'].width = img.width*0.143
        #wb.column_dimensions['B'] = Alignment(horizontal="center", vertical="center")

        wb.save(filename)


        for i in range(len(files)):
            os.remove('bad_images\\'+papka+'\\'+files[i])
        os.rmdir('bad_images\\'+papka)
    os.rmdir("bad_images")
    return 0

def delete_pictures():
    files = glob.glob(image_path+'/*')
    for f in files:
        os.remove(f)


def format_excel():
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.row_dimensions[1].height = 20
    # ffcc00 - оранжевый
    sheet['A1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['B1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['C1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['D1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['E1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['F1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['G1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['H1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['I1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['J1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['K1'].fill = PatternFill('solid', fgColor='ffcc00')

    for row in sheet[2:sheet.max_row]:  # skip the header

        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
        row[8].alignment = Alignment(horizontal="left", vertical="center")
        row[9].alignment = Alignment(horizontal="left", vertical="center")
        row[10].alignment = Alignment(horizontal="center", vertical="center")

    a = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = a
    sheet.column_dimensions['E'].width = a
    sheet.column_dimensions['F'].width = a
    sheet.column_dimensions['G'].width = a
    sheet.column_dimensions['H'].width = a

    wb.save(filename)


def main(url):
    persona = get_headers_proxy()
    html = get_html(persona, url)
    if html: parse_data(html,url)



if __name__ == '__main__':
    for url in all_url:
        main(url)
    download_img()
    img_to_excel()
    delete_pictures()
    format_excel()
    print("--- %s seconds ---" % (time.time() - start_time))