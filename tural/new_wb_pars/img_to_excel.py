import openpyxl
import os
import time
import requests
import openpyxl.drawing.image
from PIL import Image as pilIm
from openpyxl import load_workbook

file_name = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_data.xlsx'
image_path = r'C:\PyProjects\poject_1\tural\new_wb_pars\images\phone'
img_list = ['https://images.wbstatic.net/big/new/13610000/13615125-1.jpg']



def download_img(img_list):
    j = 0
    for i in img_list:
        j += 1
        img_data = requests.get(i).content
        with open(image_path+'\image_name_'+str(j)+'.jpg', 'wb') as handler:
            handler.write(img_data)


def change_size(img, shirina):
    k = img.width/shirina
    img.width = shirina
    img.height = img.height/k
    return img

def img_to_excel():
    path = os.getcwd()
    shirina = 150
    kachestvo = 45


    os.mkdir("bad_images")
    spis = os.listdir(path+'\\'+'images')

    for papka in spis:
        os.mkdir('bad_images\\'+papka)
        files = os.listdir(path+'\\'+'images\\'+papka)

    for i in files:
        try:
            image = pilIm.open(path+'\\'+'images\\'+papka+'\\'+i)
            image.save(path+'\\'+'bad_images\\'+papka+'\\'+i, quality = kachestvo)
        except:
            print('Ошибка с картинкой')
            return 1

    spis = os.listdir(path+'\\'+'bad_images')
    for papka in spis:
        files = os.listdir('bad_images\\'+papka)

        wb = load_workbook(file_name)
        sheet = wb.active

        for i in range(len(files)):
            img = openpyxl.drawing.image.Image('bad_images\\'+papka+'\\'+files[i])
            img = change_size(img, shirina)
            sheet.add_image(img, 'A%d'%(i+2))

            sheet.row_dimensions[i+2].height = img.height*0.75
        sheet.column_dimensions['A'].width = img.width*0.143
        wb.save(file_name)


        for i in range(len(files)):
            os.remove('bad_images\\'+papka+'\\'+files[i])
        os.rmdir('bad_images\\'+papka)
    os.rmdir("bad_images")
    return 0

if __name__ == '__main__':
    download_img(img_list)
    img_to_excel()
"""
import openpyxl
from openpyxl import load_workbook

file_name = r'C:\PyProjects\poject_1\tural\new_wb_pars\wildberries_data.xlsx'
img_name = r'C:\PyProjects\poject_1\tural\new_wb_pars\image_name.jpg'

wb = load_workbook(file_name)
ws = wb.worksheets[0]
img = openpyxl.drawing.image.Image(img_name)
img.anchor = 'I2'
ws.add_image(img)
wb.save(file_name)
"""