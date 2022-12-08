from bs4 import BeautifulSoup as BS
import requests
from pandas import ExcelWriter
import multiprocessing as mp
import re
import time
import string
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
start_time = time.time()

name = '11minoxidil.ru'
file_name = '11minoxidil_fucking_parsing.xlsx'

list_url = ["https://11minoxidil.ru/brends/7nutrition",
"https://11minoxidil.ru/brends/borodist",
"https://11minoxidil.ru/brends/dear-barber",
"https://11minoxidil.ru/brends/dns-roller",
"https://11minoxidil.ru/brends/epica-professional",
"https://11minoxidil.ru/brends/esthetic-house",
"https://11minoxidil.ru/brends/fallkony",
"https://11minoxidil.ru/brends/gls",
"https://11minoxidil.ru/brends/grave-before-shave",
"https://11minoxidil.ru/brends/hair-company-professional",
"https://11minoxidil.ru/brends/invit",
"https://11minoxidil.ru/brends/iron-barber",
"https://11minoxidil.ru/brends/kerastase",
"https://11minoxidil.ru/brends/lcosmetics",
"https://11minoxidil.ru/brends/loreal",
"https://11minoxidil.ru/brends/lador",
"https://11minoxidil.ru/brends/moser",
"https://11minoxidil.ru/brends/ollin-professional",
"https://11minoxidil.ru/brends/papi-and-co",
"https://11minoxidil.ru/brends/proraso",
"https://11minoxidil.ru/brends/red-star-labs",
"https://11minoxidil.ru/brends/reuzel",
"https://11minoxidil.ru/brends/seb-man",
"https://11minoxidil.ru/brends/sevich",
"https://11minoxidil.ru/brends/sim-sensitive",
"https://11minoxidil.ru/brends/solaray",
"https://11minoxidil.ru/brends/tigi",
"https://11minoxidil.ru/brends/zhangguang-fabao"]

df = pd.DataFrame(columns=['sku','name','price','obem','price_before','proizvpditel','country','nalichie', 'url'])

def parsing_kategory(url):
    link = url
    response = requests.get(link).text
    soup = BS(response, "lxml")

    try:
        soup = soup.find('div',{"class":"col-xxs-12 col-xs-12 items-container items-catalog js_fltr-items"}).find_all(class_="item-name")
    except:
        soup = soup.find('div', {"class": "col-xxs-12 col-xs-12 items-container items-catalog item-catalog-rated js_fltr-items"}).find_all(class_="item-name")

    #soup = soup.find('div', {"class": "col-xxs-12 col-xs-12 items-container items-catalog  js_fltr-items"}).find_all(class_="item-name")
    #col-xxs-12 col-xs-12 items-container items-catalog item-catalog-rated js_fltr-items

    #soup = soup.findall('div', class_="col-xxs-12 col-xs-12 items-container items-catalog item-catalog-rated js_fltr-items")
    #soup = soup.findall('div', class_="item-name")
    soup = str(soup).replace('="','99999')#.replace('">','99999')
    soup = soup.replace('">','99999')
    soup = soup.replace(' ', '99999')
    soup = soup.split('99999')

    internalLinks = []
    for i in soup:
        if len(i)>45:
            internalLinks.append(i)

    #print(internalLinks)
    #internalLinks = [a.get('href') for a in soup.find_all('a') if a.get('href') and a.get('href').startswith('/')]
    return internalLinks

def find_all_url(list_url):
    url_all = []
    for url in list_url:
        try:
            new_links = parsing_kategory(url)
            url_all = url_all + new_links
            print(url,'   OK')
        except:
            print(url,'пошел по пизде')

    url_all = ['https://11minoxidil.ru' + url_1 for url_1 in url_all]
    print(url_all)
    print(len(url_all))

    return url_all


def parsing_each_url(url):
    link_2 = url
    response = requests.get(link_2).text
    soup = BS(response, "lxml")
    box = 'item-block-right-inner'

    name = soup.find('h1').text.strip()

    price = soup.find('div', {'class': box}).find(class_="price__current").text.strip().replace(' ', '')
    price = re.sub(r'\W+', '', price)[:-1]

    sku = ''
    try:
        sku = soup.find('div', {'class': box}).find(
            class_="item-block-right__article_number item-block-right__text").text.strip().replace(' ', '')
        sku = sku.split(':')[-1]
    except:
        pass

    v = ''
    try:
        v = soup.find('div', {'class': box}).find(
            class_="item-block-right__volume item-block-right__text").text.strip().replace(' ', '')
        v = v.split(':')[-1]
    except:
        pass

    price_before = ''
    try:
        price_before = soup.find('div', {'class': box}).find(class_="price__before thro").text.strip()
        price_before = re.sub(r'\W+', '', price_before)[:-1]
    except:
        pass

    nalichie = ''
    try:
        nalichie = soup.find('div', {'class': box}).find(class_="item-block-right__action").text.strip()
        # price_before = re.sub(r'\W+', '', price_before)[:-1]
        if nalichie == 'В корзину':
            nalichie = 'В наличие'
        else:
            nalichie = 'Нет'
    except:
        pass

    proizvpditel = ''
    try:
        proizvpditel = soup.find('div', {'class': box}).find(
            class_="item-block-right__brend item-block-right__text").text
        proizvpditel = re.sub(r'\W+', ' ', proizvpditel).replace("Производитель", "").strip()
    except:
        pass

    country = soup.find('div', {'class': box}).find(class_="item-block-right__country item-block-right__text").text
    country = re.sub(r'\W+', '', country).replace("Страна", "")

    url = url

    product = {
        'sku': sku,
        'name': name,
        'price': price,
        'obem': v,
        'price_before': price_before,
        'proizvpditel': proizvpditel,
        'country': country,
        'nalichie': nalichie,
        'url': url
    }
    return product

def format_tovarn_misha(filename,shirina,color):
    df = pd.read_excel(filename)
    wb = load_workbook(filename)

    sheet = wb.active
    sheet.row_dimensions[1].height = 20

    column_list = df.columns.tolist()
    header_list = list(string.ascii_lowercase.upper())
    header_list_1 = [f'{el}1' for el in header_list]

    for i in range(df.shape[1]):
        sheet[header_list_1[i]].fill = PatternFill('solid', fgColor=color)

    for row in sheet[1:sheet.max_row]:  # skip the header
        for i in range(df.shape[1]):
            if column_list[i] not in ['Название','nazv','name','url','url_pic','Ссылка']:
                row[i].alignment = Alignment(horizontal="center", vertical="center")
            else:
                row[i].alignment = Alignment(horizontal="left", vertical="center")


    for i,j in zip(shirina,header_list):
        sheet.column_dimensions[j].width = i

    wb.save(filename)


if __name__ == '__main__':
    url_all = find_all_url(list_url)

    p = mp.Pool(processes=10)
    product = p.map(parsing_each_url,url_all)
    df = df.append(product, ignore_index=True)

    print(df)
    writer = ExcelWriter(file_name)
    df.to_excel(writer, index = False)
    writer.save()

    format_tovarn_misha(file_name, [18, 60, 10, 12, 10, 17, 10, 12, 10], 'ffcc00')
    print("--- %s seconds ---" % (time.time() - start_time))


