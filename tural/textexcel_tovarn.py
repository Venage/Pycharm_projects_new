from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )


def format_tovarn(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.row_dimensions[1].height = 20
    # ffcc00
    sheet['A1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['B1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['C1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['D1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['E1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['F1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['G1'].fill = PatternFill('solid', fgColor='ffcc00')
    sheet['H1'].fill = PatternFill('solid', fgColor='ffcc00')

    for row in sheet[1:sheet.max_row]:  # skip the header

        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(horizontal="left", vertical="center")
        row[2].alignment = Alignment(horizontal="center", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 17
    sheet.column_dimensions['F'].width = 17
    sheet.column_dimensions['G'].width = 17
    sheet.column_dimensions['H'].width = 18

    wb.save(filename)

#format_tovarn(filename)

def format_tovarn_misha(filename,shirina,color):
    import string
    import pandas as pd
    df = pd.read_excel(filename)
    wb = load_workbook(filename)

    sheet = wb.active
    sheet.row_dimensions[1].height = 20

    column_list = df.columns.tolist()
    header_list = list(string.ascii_lowercase.upper())
    header_list_1 = [f'{el}1' for el in header_list]

    for i in range(df.shape[1]):
        sheet[header_list_1[i]].fill = PatternFill('solid', fgColor=color)

    for row in sheet[1:sheet.max_row]:  # skip the header
        for i in range(df.shape[1]):
            if column_list[i] not in ['Название','nazv']:
                row[i].alignment = Alignment(horizontal="center", vertical="center")
            else:
                row[column_list.index('Название')].alignment = Alignment(horizontal="left", vertical="center")


    for i,j in zip(shirina,header_list):
        sheet.column_dimensions[j].width = i

    wb.save(filename)

filename = r"C:\Python\docs_site\tovarn_result.xlsx"
format_tovarn_misha(filename,[18,60,10,12,17,17,18],'ffcc00')